<<<<<<< HEAD
import pandas as pd
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import locale

# Tenta configurar o local para português do Brasil para obter nomes de meses e dias corretos
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    print("Locale pt_BR não encontrado, usando o padrão do sistema.")

def gerar_dashboard_oee(arquivo_entrada, ano, mes):
    """
    Gera uma planilha de Controle de OEE com uma legenda minimalista.
    """
    print(f"\n--- GERANDO RELATÓRIO PARA {mes:02d}/{ano} ---")

    # --- 1. LEITURA E PREPARAÇÃO DOS DADOS ---
    try:
        df_atividades = pd.read_csv(
            arquivo_entrada, sep=';', dtype={'circuito': str}
        )
        df_atividades['datastart'] = pd.to_datetime(df_atividades['datastart'], dayfirst=True, errors='coerce')
        df_atividades['datastop'] = pd.to_datetime(df_atividades['datastop'], dayfirst=True, errors='coerce')
        if 'status' not in df_atividades.columns:
            df_atividades['status'] = 'UP'
    except Exception as e:
        print(f"\nERRO ao ler o arquivo CSV: {e}"); return

    inicio_mes = datetime(ano, mes, 1)
    fim_mes = (inicio_mes + pd.offsets.MonthEnd(0)).to_pydatetime()
    dias_do_mes_range = pd.date_range(start=inicio_mes, end=fim_mes, freq='D')

    df_periodo = df_atividades[
        (df_atividades['datastart'].notna()) & (df_atividades['datastop'].notna()) &
        (df_atividades['datastart'] <= fim_mes) & (df_atividades['datastop'] >= inicio_mes)
    ].copy()

    circuitos_unicos = []
    if not df_atividades.empty:
        circuitos_unicos = sorted(df_atividades['circuito'].unique(), key=lambda x: int(re.search(r'\d+', x).group()))

    dados_calendario = []
    for circuito in circuitos_unicos:
        for dia in dias_do_mes_range:
            status = 'PP' if dia.weekday() >= 5 else 'SD'
            dados_calendario.append({'Circuito': circuito, 'Data': dia, 'Status': status})

    if not dados_calendario:
         print("Não foi possível gerar o calendário. Verifique o conteúdo do arquivo CSV."); return

    calendario_df = pd.DataFrame(dados_calendario).set_index(['Circuito', 'Data'])

    for _, atividade in df_periodo.iterrows():
        start_iter = max(atividade['datastart'].normalize(), pd.Timestamp(inicio_mes))
        end_iter = min(atividade['datastop'].normalize(), pd.Timestamp(fim_mes))
        status_atividade = atividade.get('status', 'UP')
        for dia in pd.date_range(start_iter, end_iter, freq='D'):
            if (atividade['circuito'], dia) in calendario_df.index:
                calendario_df.loc[(atividade['circuito'], dia), 'Status'] = status_atividade

    relatorio_detalhado_df = calendario_df.unstack(level='Data')['Status']
    relatorio_detalhado_df.columns = relatorio_detalhado_df.columns.day

    # --- 2. CRIAÇÃO E FORMATAÇÃO DO EXCEL ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Controle_OEE_{ano}_{mes:02d}"

    # --- Definição de Estilos ---
    font_title = Font(name='Calibri', size=40, bold=True, color="FFFFFF")
    font_header_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    font_legend = Font(name='Calibri', size=11, bold=True, color="000000") # Fonte preta para legenda
    font_month_year = Font(name='Calibri', size=36, bold=True, color="FFFFFF")
    font_bold_black = Font(name='Calibri', size=11, bold=True, color="000000")
    
    fill_dark_green = PatternFill(start_color="006B3D", end_color="006B3D", fill_type="solid")
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)

    fill_up = PatternFill(fill_type="solid", start_color="92D050")  # Verde
    fill_pq = PatternFill(fill_type="solid", start_color="FF0000")  # Vermelho
    fill_pp = PatternFill(fill_type="solid", start_color="9BC2E6")  # Azul
    fill_sd = PatternFill(fill_type="solid", start_color="FFEB9C")  # Amarelo

    last_day = fim_mes.day
    last_col_calendar = last_day + 1
    last_col_letter_total = get_column_letter(last_col_calendar + 8)

    # --- Montagem do Cabeçalho Principal (Linha 1) ---
    ws.merge_cells(f'A1:{last_col_letter_total}1')
    cell = ws['A1']; cell.value = "Controle de OEE - Laboratório"
    cell.fill = fill_dark_green; cell.font = font_title; cell.alignment = align_center
    ws.row_dimensions[1].height = 60

    # --- LEGENDA MINIMALISTA (Linha 3) ---
    ws.row_dimensions[3].height = 25
    
    legend_items = {
        "UP": fill_up,
        "PQ": fill_pq,
        "PP": fill_pp,
        "SD": fill_sd
    }
    
    start_col = 2 # Começa na coluna B
    for status_code, fill_color in legend_items.items():
        col_letter = get_column_letter(start_col)
        ws.column_dimensions[col_letter].width = 8
        
        cell = ws.cell(row=3, column=start_col)
        cell.value = status_code
        cell.fill = fill_color
        cell.font = font_legend
        cell.alignment = align_center
        
        start_col += 1 # Vai para a próxima coluna (C, D, E...)

    # Capacidade
    capacidade_col_start = get_column_letter(last_col_calendar + 1)
    capacidade_col_end = get_column_letter(last_col_calendar + 8)
    ws.merge_cells(f'{capacidade_col_start}3:{capacidade_col_end}3')
    cell = ws[f'{capacidade_col_start}3']; cell.value = "Capacidade de utilização: 300 circuitos"
    cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center

    # --- Mês e Ano (Linha 5) ---
    month_year_row = 5
    ws.row_dimensions[month_year_row].height = 50
    
    month_col_end = get_column_letter(last_col_calendar // 2)
    ws.merge_cells(f'A{month_year_row}:{month_col_end}{month_year_row}')
    cell = ws[f'A{month_year_row}']; cell.value = inicio_mes.strftime('%B').capitalize()
    cell.fill = fill_dark_green; cell.font = font_month_year; cell.alignment = align_center
    
    year_col_start = get_column_letter((last_col_calendar // 2) + 1)
    ws.merge_cells(f'{year_col_start}{month_year_row}:{get_column_letter(last_col_calendar)}{month_year_row}')
    cell = ws[f'{year_col_start}{month_year_row}']; cell.value = str(ano)
    cell.fill = fill_dark_green; cell.font = font_month_year; cell.alignment = align_center

    # --- Cabeçalho da Tabela (Linhas 6 e 7) ---
    header_row_dias_semana, header_row_numeros = 6, 7
    ws.row_dimensions[header_row_dias_semana].height = 25
    ws.row_dimensions[header_row_numeros].height = 25
    
    ws.merge_cells('A6:A7'); cell = ws['A6']
    cell.value = "Circuitos"; cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
    ws.column_dimensions['A'].width = 20

    dias_semana_map = ['dom', 'seg', 'ter', 'qua', 'qui', 'sex', 'sáb']
    for i, dia in enumerate(dias_do_mes_range, start=2):
        col_letter = get_column_letter(i)
        dia_da_semana_str = dias_semana_map[int(dia.strftime('%w'))]
        
        cell_dia_semana = ws.cell(row=header_row_dias_semana, column=i, value=dia_da_semana_str)
        cell_dia_semana.fill = fill_dark_green; cell_dia_semana.font = font_header_white; cell_dia_semana.alignment = align_center

        cell_dia_num = ws.cell(row=header_row_numeros, column=i, value=dia.day)
        cell_dia_num.fill = fill_dark_green; cell_dia_num.font = font_header_white; cell_dia_num.alignment = align_center
        ws.column_dimensions[col_letter].width = 5

    # --- Cabeçalho de Compilação ---
    summary_start_col = last_col_calendar + 3
    summary_end_col = summary_start_col + 3
    
    ws.merge_cells(start_row=6, start_column=summary_start_col, end_row=6, end_column=summary_end_col)
    cell = ws.cell(row=6, column=summary_start_col)
    cell.value = "Compilação"; cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center

    for i, header in enumerate(["UP", "PQ", "PP", "SD"]):
        col_idx = summary_start_col + i
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    # --- 3. INSERÇÃO DOS DADOS E FÓRMULAS ---
    current_row = header_row_numeros + 1
    if not relatorio_detalhado_df.empty:
        for circuito, data_row in relatorio_detalhado_df.iterrows():
            ws.cell(row=current_row, column=1, value=circuito).alignment = align_center
            
            for day_num, status in data_row.items():
                cell = ws.cell(row=current_row, column=day_num + 1, value=status)
                cell.alignment = align_center; cell.border = border_thin
                
            range_start = get_column_letter(2)
            range_end = get_column_letter(last_col_calendar)
            data_range = f"{range_start}{current_row}:{range_end}{current_row}"
            
            for i, status_code in enumerate(["UP", "PQ", "PP", "SD"]):
                ws.cell(row=current_row, column=summary_start_col + i, value=f'=COUNTIF({data_range}, "{status_code}")').font = font_bold_black
            
            current_row += 1

    # --- 4. FORMATAÇÃO CONDICIONAL ---
    data_range_format = f"B8:{get_column_letter(last_col_calendar)}{current_row + 5}"
    for status, fill_color in [("UP", fill_up), ("PQ", fill_pq), ("PP", fill_pp), ("SD", fill_sd)]:
        ws.conditional_formatting.add(data_range_format, CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill_color))

    ws.freeze_panes = 'B8'
    
    # --- 5. SALVAR O ARQUIVO ---
    nome_arquivo_saida = f"Dashboard_OEE_{ano}_{mes:02d}.xlsx"
    try:
        wb.save(nome_arquivo_saida)
        print(f"\n✅ Dashboard de OEE salvo com sucesso em '{nome_arquivo_saida}'!")
    except Exception as e:
        print(f"\nERRO ao salvar o arquivo Excel: {e}")

# --- BLOCO DE EXECUÇÃO INTERATIVO ---
if __name__ == "__main__":
    arquivo_de_dados = "dados_processados.csv"
    if not os.path.exists(arquivo_de_dados):
        print(f"ERRO: Arquivo de entrada '{arquivo_de_dados}' não encontrado.")
        print("Por favor, crie este arquivo com os dados de atividades para rodar o script.")
    else:
        try:
            agora = datetime.now()
            ano_padrao, mes_padrao = agora.year, agora.month

            ano_input = input(f"Digite o ano para o relatório (padrão: {ano_padrao}): ")
            ano_desejado = int(ano_input) if ano_input else ano_padrao

            while True:
                mes_input = input(f"Digite o mês para o relatório (1-12): ")
                mes_desejado = int(mes_input) if mes_input else mes_padrao
                if 1 <= mes_desejado <= 12: break
                else: print("Mês inválido. Por favor, digite um número entre 1 e 12.")
            
            gerar_dashboard_oee(arquivo_de_dados, ano_desejado, mes_desejado)
        except ValueError:
            print("Entrada inválida. Por favor, digite apenas números para ano e mês.")
        except Exception as e:
            print(f"Ocorreu um erro inesperado: {e}")
=======
import pandas as pd
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

def gerar_relatorio_oee_excel(arquivo_entrada, ano, mes):
    """
    Lê um arquivo CSV, gera um relatório de OEE diário, e trata corretamente
    atividades em andamento vs. atividades antigas sem data de fim.
    """
    print("\n--- INICIANDO SCRIPT DE RELATÓRIO OEE (LÓGICA DEFINITIVA) ---")
    
    agora = pd.Timestamp.now().tz_localize(None).normalize()

    try:
        print(f"Lendo o arquivo de dados processados: '{arquivo_entrada}'...")
        df_atividades = pd.read_csv(
            arquivo_entrada, 
            sep=';',
            dtype={'circuito': str},
            date_format='%Y-%m-%d %H:%M:%S',
            parse_dates=['datastart', 'datastop']
        )
        print(f"Sucesso! {len(df_atividades)} registros totais lidos.")
    except FileNotFoundError:
        print(f"\nERRO: Arquivo '{arquivo_entrada}' não encontrado!")
        return
    except Exception as e:
        print(f"\nERRO ao ler o arquivo CSV: {e}")
        return

    inicio_mes = pd.to_datetime(f"{ano}-{mes}-01").tz_localize(None)
    fim_mes = (inicio_mes + pd.offsets.MonthEnd(0)).tz_localize(None)
    
    df_periodo = df_atividades[
        (df_atividades['datastart'] <= fim_mes) & 
        ((df_atividades['datastop'] >= inicio_mes) | (pd.isna(df_atividades['datastop'])))
    ].copy()

    if df_periodo.empty:
        print(f"Nenhuma atividade encontrada para o período de {mes:02d}/{ano}.")
        return

    circuitos_unicos = sorted(df_periodo['circuito'].unique(), key=lambda x: int(re.search(r'\d+', x).group()))
    dias_do_mes = pd.date_range(start=inicio_mes, end=fim_mes, freq='D')
    
    print(f"\nGerando calendário de status para {len(circuitos_unicos)} circuitos potenciais...")
    
    dados_calendario = []
    for circuito in circuitos_unicos:
        for dia in dias_do_mes:
            status = 'Parada Planejada' if dia.weekday() >= 5 else 'Sem Demanda'
            dados_calendario.append({'Circuito': circuito, 'Data': dia, 'Status': status})
    
    calendario_oee_df = pd.DataFrame(dados_calendario).set_index(['Circuito', 'Data'])

    for _, atividade in df_periodo.iterrows():
        
        # --- LÓGICA DEFINITIVA PARA DATAS DE INÍCIO E FIM ---
        start_iter = max(atividade['datastart'].normalize(), inicio_mes)

        if pd.notna(atividade['datastop']):
            # CASO 1: A atividade tem data de fim. Lógica normal.
            end_iter = min(atividade['datastop'].normalize(), fim_mes)
        else:
            # CASO 2: A atividade NÃO tem data de fim.
            # Se a atividade começou ANTES do mês do relatório, é um erro de dado.
            # Contamos apenas 1 dia para não poluir o relatório.
            if atividade['datastart'] < inicio_mes:
                end_iter = start_iter 
            else:
                # Se começou NESTE mês, está em andamento. Contamos até hoje.
                end_iter = min(agora, fim_mes)
        # --- FIM DA LÓGICA DEFINITIVA ---

        if end_iter < start_iter:
            end_iter = start_iter
            
        for dia in pd.date_range(start_iter, end_iter, freq='D'):
            if (atividade['circuito'], dia) in calendario_oee_df.index:
                calendario_oee_df.loc[(atividade['circuito'], dia), 'Status'] = 'Uso Programado'

    circuitos_com_uso_real = calendario_oee_df[
        calendario_oee_df['Status'] == 'Uso Programado'
    ].index.get_level_values('Circuito').unique()

    if circuitos_com_uso_real.empty:
        print(f"Nenhum circuito teve 'Uso Programado' no período de {mes:02d}/{ano}.")
        return

    calendario_oee_df = calendario_oee_df.loc[circuitos_com_uso_real]
    print(f"Filtro aplicado! O relatório será gerado para {len(circuitos_com_uso_real)} circuitos com atividade real no mês.")
    
    relatorio_detalhado_df = calendario_oee_df.unstack(level='Data')
    relatorio_detalhado_df.columns = relatorio_detalhado_df.columns.droplevel(0)
    relatorio_detalhado_df.columns = relatorio_detalhado_df.columns.strftime('%d/%m/%Y')
    relatorio_detalhado_df.index.name = 'Circuito'

    resumo_oee = calendario_oee_df.groupby('Circuito')['Status'].value_counts().unstack(fill_value=0)
    
    for status_col in ['Uso Programado', 'Sem Demanda', 'Parada Planejada']:
        if status_col not in resumo_oee.columns:
            resumo_oee[status_col] = 0
            
    print("\n--- Relatório de OEE por Circuito (Contagem de Dias) ---")
    print(resumo_oee[['Uso Programado', 'Sem Demanda', 'Parada Planejada']].to_string())
    
    nome_arquivo_saida = f"relatorio_oee_visual_{ano}-{mes:02d}.xlsx"
    print(f"\nSalvando e formatando o relatório em '{nome_arquivo_saida}'...")

    with pd.ExcelWriter(nome_arquivo_saida, engine='openpyxl') as writer:
        relatorio_detalhado_df.to_excel(writer, sheet_name='Relatorio_OEE_Diario')
        resumo_oee[['Uso Programado', 'Sem Demanda', 'Parada Planejada']].to_excel(writer, sheet_name='Resumo_Contagem')
        
        ws_detalhado = writer.sheets['Relatorio_OEE_Diario']
        fill_up = PatternFill(fill_type="solid", start_color="C6EFCE")
        fill_sd = PatternFill(fill_type="solid", start_color="FFEB9C")
        fill_pp = PatternFill(fill_type="solid", start_color="FFC7CE")

        max_row = ws_detalhado.max_row
        max_col_letter = get_column_letter(ws_detalhado.max_column)
        cell_range = f"B2:{max_col_letter}{max_row}"

        ws_detalhado.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Uso Programado"'], fill=fill_up))
        ws_detalhado.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Sem Demanda"'], fill=fill_sd))
        ws_detalhado.conditional_formatting.add(cell_range, CellIsRule(operator='equal', formula=['"Parada Planejada"'], fill=fill_pp))

        ws_detalhado.column_dimensions['A'].width = 15
        for col_idx in range(2, ws_detalhado.max_column + 1):
            ws_detalhado.column_dimensions[get_column_letter(col_idx)].width = 15
            
        ws_resumo = writer.sheets['Resumo_Contagem']
        ws_resumo.column_dimensions['A'].width = 15
        for col_letter in ['B', 'C', 'D']:
            ws_resumo.column_dimensions[col_letter].width = 20

    print(f"\n✅ Relatório OEE visual salvo com sucesso!")

if __name__ == "__main__":
    arquivo_de_dados = "dados_processados.csv"
    if not os.path.exists(arquivo_de_dados):
        print(f"ERRO: Arquivo de entrada '{arquivo_de_dados}' não encontrado.")
    else:
        try:
            ano_atual = datetime.now().year
            mes_atual = datetime.now().month
            ano_desejado_input = input(f"Digite o ano (padrão: {ano_atual}): ")
            ano_desejado = int(ano_desejado_input) if ano_desejado_input else ano_atual
            mes_desejado_input = input(f"Digite o mês (padrão: {mes_atual}): ")
            mes_desejado = int(mes_desejado_input) if mes_desejado_input else mes_atual
            if not (1 <= mes_desejado <= 12):
                raise ValueError("Mês inválido.")
        except ValueError as e:
            print(f"Entrada inválida. {e}.")
        else:
            gerar_relatorio_oee_excel(arquivo_de_dados, ano_desejado, mes_desejado)
>>>>>>> b168e54c4147bafe255aaec9747ff4a520d69bd9
