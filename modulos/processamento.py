import pandas as pd
import re
import os
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import locale

try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    pass

def interpretar_data_flexivel(string_data, dayfirst=True):
    if pd.isna(string_data) or not isinstance(string_data, str):
        return pd.NaT
    string_data = string_data.strip()
    if re.search(r'\b(am|pm)\b', string_data, re.IGNORECASE):
        return pd.NaT
    return pd.to_datetime(string_data, dayfirst=dayfirst, errors='coerce')

def limpar_dados_brutos(lista_arquivos_path, arquivo_saida_path, formato_data="dd/mm/aaaa"):
    all_lines = []
    for nome_arquivo in lista_arquivos_path:
        try:
            try:
                with open(nome_arquivo, 'r', encoding='utf-8') as f: lines = f.readlines()
            except UnicodeDecodeError:
                with open(nome_arquivo, 'r', encoding='latin-1') as f: lines = f.readlines()
            all_lines.extend(lines)
        except FileNotFoundError: continue
    if not all_lines: return (False, [])
    
    conteudo_total = "\n".join(all_lines)
    circuit_pattern = re.compile(r"Circuit\d+", re.IGNORECASE)
    
    datetime_pattern = re.compile(r"\b\d{1,2}[/-]\d{1,2}[/-]\d{2,4}(?:\s+\d{1,2}:\d{2}(?::\d{2})?)?\b")
    
    data_list = []
    circuit_matches = list(circuit_pattern.finditer(conteudo_total))

    for i, match in enumerate(circuit_matches):
        circuito = match.group(0)
        start_pos = match.end()
        end_pos = circuit_matches[i+1].start() if i + 1 < len(circuit_matches) else len(conteudo_total)
        data_blob = conteudo_total[start_pos:end_pos]
        found_datetimes = datetime_pattern.findall(data_blob)
        datastart_str = found_datetimes[0] if len(found_datetimes) >= 1 else None
        datastop_str = found_datetimes[1] if len(found_datetimes) >= 2 else None
        if datastart_str: data_list.append({'circuito': circuito, 'datastart': datastart_str, 'datastop': datastop_str})
    
    if not data_list: return (False, [])
    df = pd.DataFrame(data_list)
    
    dayfirst_bool = (formato_data == "dd/mm/aaaa")
    
    df['datastart'] = df['datastart'].apply(interpretar_data_flexivel, dayfirst=dayfirst_bool)
    df['datastop'] = df['datastop'].apply(interpretar_data_flexivel, dayfirst=dayfirst_bool)
    
    df.dropna(subset=['datastart'], inplace=True)
    if df.empty: return (False, [])
        
    df['circuito_num'] = df['circuito'].str.extract(r'(\d+)').astype(int)
    df.sort_values(by=['circuito_num', 'datastart'], ascending=[True, False], inplace=True)
    df.drop(columns=['circuito_num'], inplace=True)
    df.reset_index(drop=True, inplace=True)
    
    df.to_csv(arquivo_saida_path, index=False, sep=';', date_format='%d/%m/%Y %H:%M:%S', na_rep='')
    
    circuitos_unicos = sorted(df['circuito'].unique(), key=lambda x: int(re.search(r'\d+', x).group()))
    return (True, circuitos_unicos)

def salvar_historico_csv(ano, mes, sumario, pasta_saida):
    caminho_csv = os.path.join(pasta_saida, 'historico_oee.csv')
    colunas = ['ano', 'mes', 'disponibilidade', 'performance', 'qualidade', 'oee_final']
    nova_linha = {
        'ano': ano,
        'mes': mes,
        'disponibilidade': sumario.get('Disponibilidade', 0),
        'performance': sumario.get('Performance', 0),
        'qualidade': sumario.get('Qualidade', 0),
        'oee_final': sumario.get('OEE', 0)
    }
    df_novo = pd.DataFrame([nova_linha])
    if os.path.exists(caminho_csv):
        df_historico = pd.read_csv(caminho_csv)
        filtro = (df_historico['ano'] == ano) & (df_historico['mes'] == mes)
        if not df_historico[filtro].empty:
            df_historico = df_historico[~filtro]
        df_final = pd.concat([df_historico, df_novo], ignore_index=True)
    else:
        df_final = df_novo
    df_final = df_final.sort_values(by=['ano', 'mes'])
    df_final.to_csv(caminho_csv, index=False)

def gerar_dashboard_oee(arquivo_entrada_path, arquivo_saida_folder, ano, mes, capacidade_total=300, regras_de_force=None, min_dias_up=1, aplicar_min_dias_up=True, ensaios_executados=None, ensaios_solicitados=None, relatorios_no_prazo=None, relatorios_emitidos=None):
    if regras_de_force is None:
        regras_de_force = {}
    
    try:
        df_atividades = pd.read_csv(arquivo_entrada_path, sep=';', dtype={'circuito': str}, na_values='')
        df_atividades['datastart'] = pd.to_datetime(df_atividades['datastart'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
        df_atividades['datastop'] = pd.to_datetime(df_atividades['datastop'], format='%d/%m/%Y %H:%M:%S', errors='coerce')
        if 'status' not in df_atividades.columns:
            df_atividades['status'] = 'UP'
    except Exception as e:
        print(f"Erro ao ler CSV: {e}")
        return None

    inicio_mes = datetime(ano, mes, 1)
    fim_mes = (inicio_mes + pd.offsets.MonthEnd(0)).to_pydatetime()
    fim_do_ultimo_dia = fim_mes.replace(hour=23, minute=59, second=59)

    df_atividades['datastop'] = df_atividades['datastop'].fillna(fim_do_ultimo_dia)
    df_atividades.dropna(subset=['datastart'], inplace=True)

    dias_do_mes_range = pd.date_range(start=inicio_mes, end=fim_mes, freq='D')
    
    todos_circuitos_no_arquivo = set(df_atividades['circuito'].unique())
    
    circuitos_up_force = regras_de_force.get('circuitos_up', [])
    circuitos_pq_force = regras_de_force.get('circuitos_pq', [])
    circuitos_vazio_force = regras_de_force.get('circuitos_vazio', [])
    
    circuitos_para_processar = (
        todos_circuitos_no_arquivo.union(set(circuitos_up_force))
        .union(set(circuitos_pq_force))
    ).difference(set(circuitos_vazio_force))
    
    def custom_sort_key(index):
        numeric_parts = index.str.extract(r'(\d+)').iloc[:, 0].fillna('9999').astype(int)
        numeric_parts[index == 'iDevice'] = -1 
        return numeric_parts
    
    circuitos_para_processar.discard('iDevice')
    
    circuitos_para_processar = sorted(list(circuitos_para_processar), key=lambda x: int(re.search(r'\d+', x).group()))

    idevice_data = {}
    for dia in dias_do_mes_range:
        status = 'UP' if dia.weekday() < 5 else 'PP'
        idevice_data[dia.day] = status
    idevice_series = pd.Series(idevice_data, name='iDevice')
    
    if not circuitos_para_processar:
        calendario_df = pd.DataFrame()
    else:
        dados_calendario = []
        for circuito in circuitos_para_processar:
            for dia in dias_do_mes_range:
                status = 'PP' if dia.weekday() >= 5 else 'SD'
                dados_calendario.append({'Circuito': circuito, 'Data': dia, 'Status': status})

        calendario_df = pd.DataFrame(dados_calendario).set_index(['Circuito', 'Data'])
    
    if not calendario_df.empty:
        df_periodo = df_atividades[df_atividades['circuito'].isin(circuitos_para_processar)].copy()
    
        for _, atividade in df_periodo.iterrows():
            start_day = max(atividade['datastart'].normalize(), pd.Timestamp(inicio_mes))
            end_day = min(atividade['datastop'].normalize(), pd.Timestamp(fim_mes))
            status_atividade = atividade.get('status', 'UP')
            
            for dia in pd.date_range(start=start_day, end=end_day, freq='D'):
                if (atividade['circuito'], dia) in calendario_df.index:
                    calendario_df.loc[(atividade['circuito'], dia), 'Status'] = status_atividade

        tipo_up_force = regras_de_force.get('tipo_up')
        for circuito in circuitos_up_force:
            if circuito in calendario_df.index.get_level_values('Circuito'):
                if tipo_up_force == "Forçar 100% UP":
                    calendario_df.loc[circuito, 'Status'] = 'UP'
                elif tipo_up_force == "Forçar Semana Padrão (Seg-Sex UP)":
                    for dia in dias_do_mes_range:
                        if (circuito, dia) in calendario_df.index:
                            status_forcado = 'UP' if dia.weekday() < 5 else 'PP'
                            calendario_df.loc[(circuito, dia), 'Status'] = status_forcado
        
        for circuito in circuitos_pq_force:
            if circuito in calendario_df.index.get_level_values('Circuito'):
                calendario_df.loc[circuito, 'Status'] = 'PQ'

    if not calendario_df.empty:
        relatorio_detalhado_df = calendario_df.unstack(level='Data')['Status']
        relatorio_detalhado_df.columns = relatorio_detalhado_df.columns.day
    else:
        relatorio_detalhado_df = pd.DataFrame(columns=[d.day for d in dias_do_mes_range])

    indices_apenas_sd_pp = []
    for index, row in relatorio_detalhado_df.iterrows():
        valores_unicos = row.dropna().unique()
        if len(valores_unicos) > 0 and all(val in ['PP', 'SD'] for val in valores_unicos):
            indices_apenas_sd_pp.append(index)

    df_para_calculo = relatorio_detalhado_df.drop(index=indices_apenas_sd_pp, errors='ignore')

    df_para_calculo.loc['iDevice'] = idevice_series
    
    if aplicar_min_dias_up:
        indices_para_remover = []
        for index, row in df_para_calculo.iterrows():
            if index in circuitos_up_force or index in circuitos_pq_force or index == 'iDevice':
                continue
            dias_up_reais = (row == 'UP').sum()
            if dias_up_reais < min_dias_up:
                indices_para_remover.append(index)
        df_para_calculo.drop(indices_para_remover, inplace=True)

    df_para_calculo = df_para_calculo.loc[(df_para_calculo != '').any(axis=1) & (df_para_calculo.notna().any(axis=1))]

    df_sd_pp_vazio = pd.DataFrame(index=indices_apenas_sd_pp, columns=relatorio_detalhado_df.columns).fillna('')
    circuitos_vazio_df = pd.DataFrame(index=circuitos_vazio_force, columns=relatorio_detalhado_df.columns).fillna('')
    df_vazios_total = pd.concat([df_sd_pp_vazio, circuitos_vazio_df])

    if not df_vazios_total.empty:
        relatorio_detalhado_df = pd.concat([df_para_calculo, df_vazios_total]).sort_index(key=custom_sort_key)
    else:
        relatorio_detalhado_df = df_para_calculo.sort_index(key=custom_sort_key)
    
    circuitos_usados_df = df_para_calculo
    circuitos_usados_count = len(circuitos_usados_df)

    sumario_ui = {'totais': {}, 'medias': {}}
    
    medias = {}
    if circuitos_usados_count > 0:
        totais = {
            'UP': (circuitos_usados_df == 'UP').sum().sum(),
            'PQ': (circuitos_usados_df == 'PQ').sum().sum(),
            'PP': (circuitos_usados_df == 'PP').sum().sum(),
            'SD': (circuitos_usados_df == 'SD').sum().sum()
        }
        sumario_ui['totais'] = totais

        medias = {status: total / circuitos_usados_count for status, total in totais.items()}
        sumario_ui['medias'] = {status: math.ceil(total) for status, total in medias.items()}
        
    total_dias_mes = len(dias_do_mes_range)

    tempo_disponivel = total_dias_mes - medias.get('PP', 0) - medias.get('SD', 0)
    tempo_real_op = medias.get('UP', 0) - medias.get('PQ', 0) - medias.get('SD', 0)
    disponibilidade = tempo_real_op / tempo_disponivel if tempo_disponivel > 0 else 0
    
    performance = ensaios_executados / ensaios_solicitados if ensaios_solicitados and ensaios_solicitados > 0 else 0
    
    qualidade = relatorios_no_prazo / relatorios_emitidos if relatorios_emitidos and relatorios_emitidos > 0 else 0
    
    oee = disponibilidade * performance * qualidade
    
    sumario_ui['tempo_disponivel'] = tempo_disponivel
    sumario_ui['tempo_real_op'] = tempo_real_op
    sumario_ui['ensaios_executados'] = ensaios_executados
    sumario_ui['ensaios_solicitados'] = ensaios_solicitados
    sumario_ui['relatorios_no_prazo'] = relatorios_no_prazo
    sumario_ui['relatorios_emitidos'] = relatorios_emitidos
    sumario_ui['Disponibilidade'] = round(disponibilidade * 100, 2)
    sumario_ui['Performance'] = round(performance * 100, 2)
    sumario_ui['Qualidade'] = round(qualidade * 100, 2)
    sumario_ui['OEE'] = round(oee * 100, 2)

    sumario_ui['circuitos_usados'] = circuitos_usados_count
    sumario_ui['circuitos_total'] = capacidade_total
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Controle_OEE_{ano}_{mes:02d}"
    
    font_title = Font(name='Calibri', size=40, bold=True, color="FFFFFF")
    font_header_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    font_legend = Font(name='Calibri', size=11, bold=True, color="000000")
    font_month_year = Font(name='Calibri', size=36, bold=True, color="FFFFFF")
    font_bold_black = Font(name='Calibri', size=11, bold=True, color="000000")
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_thin_side = Side(border_style="thin", color="D9D9D9")
    border_thin_all = Border(left=border_thin_side, right=border_thin_side, top=border_thin_side, bottom=border_thin_side)
    
    fill_dark_green = PatternFill(start_color="006B3D", end_color="006B3D", fill_type="solid")
    fill_up = PatternFill(fill_type="solid", start_color="92D050")
    fill_pq = PatternFill(fill_type="solid", start_color="FF0000")
    fill_pp = PatternFill(fill_type="solid", start_color="9BC2E6")
    fill_sd = PatternFill(fill_type="solid", start_color="FFEB9C")
    
    fill_oee_header = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_oee_subheader = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    font_oee_header = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    font_oee_sub = Font(name='Calibri', size=11, bold=True, color="000000")
    
    last_day = fim_mes.day
    
    col_calendar_start = 2
    col_calendar_end = col_calendar_start + last_day - 1
    summary_start_col = col_calendar_end + 2
    
    oee_table_start_col = summary_start_col + 5
    col_oee_label = oee_table_start_col
    col_oee_value = oee_table_start_col + 1
    col_result_label = oee_table_start_col + 3
    col_result_value = oee_table_start_col + 4
    
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions[get_column_letter(col_oee_label)].width = 25
    ws.column_dimensions[get_column_letter(col_oee_value)].width = 15
    ws.column_dimensions[get_column_letter(col_oee_value + 1)].width = 2
    ws.column_dimensions[get_column_letter(col_result_label)].width = 25
    ws.column_dimensions[get_column_letter(col_result_value)].width = 15

    last_col_letter_total = get_column_letter(col_result_value)
    ws.merge_cells(f'A1:{last_col_letter_total}1')
    cell = ws['A1']; cell.value = "Controle de OEE - Laboratório"
    cell.fill = fill_dark_green; cell.font = font_title; cell.alignment = align_center
    ws.row_dimensions[1].height = 60
    
    ws.row_dimensions[3].height = 25
    legend_items = {"UP": fill_up, "PQ": fill_pq, "PP": fill_pp, "SD": fill_sd}
    start_col_legend = 2
    for status_code, fill_color in legend_items.items():
        cell = ws.cell(row=3, column=start_col_legend)
        cell.value = status_code; cell.fill = fill_color; cell.font = font_legend; cell.alignment = align_center
        ws.column_dimensions[get_column_letter(start_col_legend)].width = 8
        start_col_legend += 1
        
    capacidade_col_start = summary_start_col
    capacidade_col_end = summary_start_col + 7
    ws.merge_cells(start_row=3, start_column=capacidade_col_start, end_row=3, end_column=capacidade_col_end)
    cell = ws.cell(row=3, column=capacidade_col_start); cell.value = f"Capacidade de utilização: {capacidade_total} circuitos"
    cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
    
    month_year_row = 5
    ws.row_dimensions[month_year_row].height = 50
    month_col_end = col_calendar_start + (last_day // 2)
    ws.merge_cells(start_row=month_year_row, start_column=col_calendar_start, end_row=month_year_row, end_column=month_col_end)
    cell = ws.cell(row=month_year_row, column=col_calendar_start); cell.value = inicio_mes.strftime('%B').capitalize()
    cell.fill = fill_dark_green; cell.font = font_month_year; cell.alignment = align_center
    
    year_col_start = month_col_end + 1
    ws.merge_cells(start_row=month_year_row, start_column=year_col_start, end_row=month_year_row, end_column=col_calendar_end)
    cell = ws.cell(row=month_year_row, column=year_col_start); cell.value = str(ano)
    cell.fill = fill_dark_green; cell.font = font_month_year; cell.alignment = align_center
    
    header_row_dias_semana, header_row_numeros = 6, 7
    ws.row_dimensions[header_row_dias_semana].height = 25
    ws.row_dimensions[header_row_numeros].height = 25
    
    ws.merge_cells(start_row=6, start_column=1, end_row=7, end_column=1)
    cell = ws.cell(row=6, column=1); cell.value = "Circuitos"
    cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
    
    dias_semana_map = ['dom', 'seg', 'ter', 'qua', 'qui', 'sex', 'sáb']
    for i, dia in enumerate(dias_do_mes_range, start=col_calendar_start):
        col_letter = get_column_letter(i)
        cell_semana = ws.cell(row=header_row_dias_semana, column=i, value=dias_semana_map[int(dia.strftime('%w'))])
        cell_semana.fill = fill_dark_green; cell_semana.font = font_header_white; cell_semana.alignment = align_center
        cell_num = ws.cell(row=header_row_numeros, column=i, value=dia.day)
        cell_num.fill = fill_dark_green; cell_num.font = font_header_white; cell_num.alignment = align_center
        ws.column_dimensions[col_letter].width = 5
        
    ws.merge_cells(start_row=6, start_column=summary_start_col, end_row=6, end_column=summary_start_col + 3)
    cell = ws.cell(row=6, column=summary_start_col)
    cell.value = "Compilação"; cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
    
    for i, header in enumerate(["UP", "PQ", "PP", "SD"]):
        col_idx = summary_start_col + i
        cell = ws.cell(row=7, column=col_idx, value=header)
        cell.fill = fill_dark_green; cell.font = font_header_white; cell.alignment = align_center
        ws.column_dimensions[get_column_letter(col_idx)].width = 7
    
    start_data_row = header_row_numeros + 1
    current_row = start_data_row
    if not relatorio_detalhado_df.empty:
        for circuito, data_row in relatorio_detalhado_df.iterrows():
            ws.cell(row=current_row, column=1, value=circuito).alignment = align_center
            for day_num, status in data_row.items():
                cell = ws.cell(row=current_row, column=col_calendar_start + day_num -1, value=status)
                cell.alignment = align_center; cell.border = border_thin_all
            data_range = f"{get_column_letter(col_calendar_start)}{current_row}:{get_column_letter(col_calendar_end)}{current_row}"
            for i, status_code in enumerate(["UP", "PQ", "PP", "SD"]):
                ws.cell(row=current_row, column=summary_start_col + i, value=f'=COUNTIF({data_range}, "{status_code}")').font = font_bold_black
            current_row += 1

    media_row = current_row + 1
    if circuitos_usados_count > 0:
        label_media_cell = ws.cell(row=media_row, column=1, value="MÉDIA")
        label_media_cell.font = font_bold_black
        label_media_cell.fill = fill_oee_subheader
        label_media_cell.alignment = align_center
        style_media = NamedStyle(name='media_style', font=font_bold_black, number_format='0')
        for i in range(4):
            col_idx = summary_start_col + i
            col_letter = get_column_letter(col_idx)
            sum_range = f"{col_letter}{start_data_row}:{col_letter}{current_row - 1}"
            formula_media = f"=ROUNDUP(SUM({sum_range})/{circuitos_usados_count}, 0)"
            media_cell = ws.cell(row=media_row, column=col_idx, value=formula_media)
            media_cell.style = style_media
            media_cell.border = border_thin_all
            media_cell.fill = fill_oee_subheader

    oee_start_row = 9
    col_label = oee_table_start_col
    col_value = oee_table_start_col + 1
    col_result_label = oee_table_start_col + 3
    col_result_value = oee_table_start_col + 4
    
    ws.merge_cells(start_row=oee_start_row, start_column=col_label, end_row=oee_start_row, end_column=col_value)
    cell = ws.cell(row=oee_start_row, column=col_label, value="Valores para o Cálculo do OEE")
    cell.fill = fill_oee_header
    cell.font = font_oee_header
    cell.alignment = align_center
    cell.border = border_thin_all
    
    row_tempo_disp = oee_start_row + 1
    cell = ws.cell(row=row_tempo_disp, column=col_label, value="Tempo Disponível (dias)")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_tempo_disp, column=col_value, value=sumario_ui.get('tempo_disponivel', 0)).number_format = '0.00'
    ws.cell(row=row_tempo_disp, column=col_value).border = border_thin_all
    
    row_tempo_real = oee_start_row + 2
    cell = ws.cell(row=row_tempo_real, column=col_label, value="Tempo Real Utilizado (dias)")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_tempo_real, column=col_value, value=sumario_ui.get('tempo_real_op', 0)).number_format = '0.00'
    ws.cell(row=row_tempo_real, column=col_value).border = border_thin_all
    
    row_ensaios_sol = oee_start_row + 3
    cell = ws.cell(row=row_ensaios_sol, column=col_label, value="Ensaios Solicitados")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_ensaios_sol, column=col_value, value=sumario_ui.get('ensaios_solicitados', 0)).border = border_thin_all
    
    row_ensaios_exec = oee_start_row + 4
    cell = ws.cell(row=row_ensaios_exec, column=col_label, value="Ensaios Executados")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_ensaios_exec, column=col_value, value=sumario_ui.get('ensaios_executados', 0)).border = border_thin_all
    
    row_rel_emit = oee_start_row + 5
    cell = ws.cell(row=row_rel_emit, column=col_label, value="Relatórios Emitidos")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_rel_emit, column=col_value, value=sumario_ui.get('relatorios_emitidos', 0)).border = border_thin_all
    
    row_rel_prazo = oee_start_row + 6
    cell = ws.cell(row=row_rel_prazo, column=col_label, value="Relatórios no Prazo")
    cell.fill = fill_oee_subheader
    cell.font = font_oee_sub
    cell.border = border_thin_all
    ws.cell(row=row_rel_prazo, column=col_value, value=sumario_ui.get('relatorios_no_prazo', 0)).border = border_thin_all
    
    ws.merge_cells(start_row=oee_start_row, start_column=col_result_label, end_row=oee_start_row, end_column=col_result_value)
    cell = ws.cell(row=oee_start_row, column=col_result_label, value="Resultados do OEE")
    cell.fill = fill_oee_header
    cell.font = font_oee_header
    cell.alignment = align_center
    cell.border = border_thin_all

    row_disp_final = oee_start_row + 1
    cell = ws.cell(row=row_disp_final, column=col_result_label, value="Disponibilidade")
    cell.font = font_bold_black
    cell.border = border_thin_all
    cell = ws.cell(row=row_disp_final, column=col_result_value, value=f"={get_column_letter(col_value)}{row_tempo_real}/{get_column_letter(col_value)}{row_tempo_disp}")
    cell.style = 'Percent'
    cell.border = border_thin_all
    
    row_perf_final = oee_start_row + 2
    cell = ws.cell(row=row_perf_final, column=col_result_label, value="Performance")
    cell.font = font_bold_black
    cell.border = border_thin_all
    cell = ws.cell(row=row_perf_final, column=col_result_value, value=f"={get_column_letter(col_value)}{row_ensaios_exec}/{get_column_letter(col_value)}{row_ensaios_sol}")
    cell.style = 'Percent'
    cell.border = border_thin_all

    row_qual_final = oee_start_row + 3
    cell = ws.cell(row=row_qual_final, column=col_result_label, value="Qualidade")
    cell.font = font_bold_black
    cell.border = border_thin_all
    cell = ws.cell(row=row_qual_final, column=col_result_value, value=f"={get_column_letter(col_value)}{row_rel_prazo}/{get_column_letter(col_value)}{row_rel_emit}")
    cell.style = 'Percent'
    cell.border = border_thin_all
    
    row_oee_final = oee_start_row + 5
    
    cell = ws.cell(row=row_oee_final, column=col_result_label, value="OEE Final")
    cell.font = Font(name='Calibri', size=11, bold=True, color="FF0000")
    cell.border = border_thin_all
    
    cell = ws.cell(row=row_oee_final, column=col_result_value, value=f"={get_column_letter(col_result_value)}{row_disp_final}*{get_column_letter(col_result_value)}{row_perf_final}*{get_column_letter(col_result_value)}{row_qual_final}")
    cell.style = 'Percent'
    cell.border = border_thin_all
    
    data_range_format = f"{get_column_letter(col_calendar_start)}{start_data_row}:{get_column_letter(col_calendar_end)}{current_row + 5}"
    for status, fill_color in [("UP", fill_up), ("PQ", fill_pq), ("PP", fill_pp), ("SD", fill_sd)]:
        ws.conditional_formatting.add(data_range_format, CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill_color))
    
    ws.freeze_panes = f'{get_column_letter(col_calendar_start)}{header_row_numeros + 1}'
    
    nome_arquivo_saida = f"Excel_OEE_{ano}_{mes:02d}.xlsx"
    caminho_completo_saida = os.path.join(arquivo_saida_folder, nome_arquivo_saida)

    try:
        wb.save(caminho_completo_saida)
    except Exception as e:
        print(f"Erro ao salvar o excel: {e}")
        return None
    
    salvar_historico_csv(ano, mes, sumario_ui, arquivo_saida_folder)

    resultados = {
        'caminho_excel': caminho_completo_saida,
        'sumario': sumario_ui,
        'df_preview': relatorio_detalhado_df
    }
    return resultados